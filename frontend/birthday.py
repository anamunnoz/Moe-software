import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QMessageBox, QGroupBox, QProgressBar, QFrame, QComboBox, QCheckBox,
    QScrollArea, QSizePolicy, QApplication
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QPixmap, QClipboard
from utils import http_get, http_put
from urls import API_URL_ORDERS, API_URL_CLIENTES
from datetime import datetime 
from openpyxl.utils import get_column_letter


class BirthdayTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_birthday_tab()

    def _setup_birthday_tab(self):
        main_layout = QVBoxLayout(self)
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)

        # ---------------- ENCABEZADO ----------------
        header_layout = QHBoxLayout()
        header_icon = QLabel()
        header_icon.setPixmap(
            QPixmap("icons/birthday.png").scaled(50, 50, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        )
        header_title = QLabel("Cumpleaños de Hoy")
        header_title.setStyleSheet("""
            font-size: 28px;
            font-weight: bold;
            color: #2c3e50;
        """)
        
        today_label = QLabel(f"{datetime.now().strftime('%d/%m/%Y')}")
        today_label.setStyleSheet("""
            font-size: 16px;
            color: #7f8c8d;
            font-weight: 500;
        """)

        header_layout.addWidget(header_icon)
        header_layout.addWidget(header_title)
        header_layout.addStretch()
        header_layout.addWidget(today_label)
        
        main_layout.addLayout(header_layout)

        # ---------------- ÁREA DESPLAZABLE ----------------
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background-color: #f8f9fa;
                width: 10px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background-color: #bdc3c7;
                border-radius: 5px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #95a5a6;
            }
        """)

        # Widget contenedor para las tarjetas
        self.cards_container = QWidget()
        self.cards_layout = QVBoxLayout(self.cards_container)
        self.cards_layout.setSpacing(20)
        self.cards_layout.setAlignment(Qt.AlignTop)
        self.cards_layout.setContentsMargins(10, 10, 10, 10)

        scroll_area.setWidget(self.cards_container)
        main_layout.addWidget(scroll_area)

        # ---------------- BOTÓN ACTUALIZAR ----------------
        refresh_btn = QPushButton("🔄 Actualizar Lista")
        refresh_btn.setFixedHeight(40)
        refresh_btn.clicked.connect(self._load_birthday_clients)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-weight: 600;
                padding: 8px 16px;
                border-radius: 6px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        main_layout.addWidget(refresh_btn, alignment=Qt.AlignCenter)

        # Cargar clientes automáticamente al abrir la pestaña
        self._load_birthday_clients()

    def _load_birthday_clients(self):
        # Limpiar tarjetas existentes
        for i in reversed(range(self.cards_layout.count())): 
            widget = self.cards_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()

        try:
            print("🎂 Cargando clientes con cumpleaños...")
            response = http_get(API_URL_CLIENTES)
            
            if not response or response.status_code != 200:
                self._show_no_clients_message("Error al cargar clientes")
                return

            clients = response.json()
            print(f"📋 Total de clientes obtenidos: {len(clients)}")

            # Obtener fecha actual
            today = datetime.now()
            current_day = today.day
            current_month = today.month

            birthday_clients = []
            
            for client in clients:
                identity = client.get('identity', '')
                if len(identity) >= 6:
                    try:
                        # Extraer día y mes (posiciones 3-6: índice 2-5)
                        day = int(identity[4:6])
                        month = int(identity[2:4])

                        print("day", day)
                        print('month', month)
                        
                        if day == current_day and month == current_month:
                            birthday_clients.append(client)
                    except (ValueError, IndexError):
                        continue

            print(f"🎉 Clientes que cumplen hoy: {len(birthday_clients)}")
            
            if not birthday_clients:
                self._show_no_clients_message("🎉 ¡No hay cumpleaños hoy!")
                return

            # Crear tarjetas para cada cliente
            for client in birthday_clients:
                self._create_client_card(client)

        except Exception as e:
            print(f"❌ Error cargando cumpleaños: {e}")
            self._show_no_clients_message("Error al cargar cumpleaños")

    def _show_no_clients_message(self, message):
        no_clients_label = QLabel(message)
        no_clients_label.setStyleSheet("""
            font-size: 18px;
            color: #7f8c8d;
            font-weight: 500;
            padding: 40px;
            text-align: center;
        """)
        no_clients_label.setAlignment(Qt.AlignCenter)
        self.cards_layout.addWidget(no_clients_label)

    def _create_client_card(self, client):
        # Crear tarjeta
        card = QWidget()
        card.setStyleSheet("""
            QWidget {
                background-color: #ffffff;
                border: 2px solid #e74c3c;
                border-radius: 12px;
                padding: 0px;
            }
        """)
        card.setFixedHeight(120)
        
        card_layout = QHBoxLayout(card)
        card_layout.setContentsMargins(20, 15, 20, 15)
        card_layout.setSpacing(20)

        # ---------------- INFORMACIÓN DEL CLIENTE ----------------
        info_widget = QWidget()
        info_layout = QVBoxLayout(info_widget)
        info_layout.setContentsMargins(0, 0, 0, 0)
        info_layout.setSpacing(8)

        # Nombre del cliente (grande)
        name_label = QLabel(client.get('name', 'Cliente'))
        name_label.setStyleSheet("""
            font-size: 20px;
            font-weight: bold;
            color: #2c3e50;
        """)
        name_label.setWordWrap(True)

        # Información adicional (pequeña)
        details_layout = QHBoxLayout()
        details_layout.setSpacing(15)

        identity = client.get('identity', '')
        phone = client.get('phone_number', 'No disponible')
        
        identity_label = QLabel(f"🆔 {identity}")
        identity_label.setStyleSheet("font-size: 13px; color: #7f8c8d;")

        phone_label = QLabel(f"📞 {phone}")
        phone_label.setStyleSheet("font-size: 13px; color: #7f8c8d;")

        details_layout.addWidget(identity_label)
        details_layout.addWidget(phone_label)
        details_layout.addStretch()

        info_layout.addWidget(name_label)
        info_layout.addLayout(details_layout)

        # ---------------- BOTÓN COPIAR ----------------
        copy_btn = QPushButton("📋 Copiar Mensaje")
        copy_btn.setFixedSize(120, 40)
        copy_btn.clicked.connect(lambda: self._copy_birthday_message(client))
        copy_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: 600;
                padding: 8px 12px;
                border-radius: 6px;
                border: none;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
            QPushButton:pressed {
                background-color: #1e8449;
            }
        """)

        card_layout.addWidget(info_widget)
        card_layout.addWidget(copy_btn)

        self.cards_layout.addWidget(card)

    def _copy_birthday_message(self, client):
        client_name = client.get('name', 'Cliente')
        message = f"""🎉 ¡FELIZ CUMPLEAÑOS {client_name.upper()}! 🎉
        ¡En nombre de todo nuestro equipo queremos desearte un maravilloso día lleno de alegría y bendiciones! 🎂🎁
        Para hacer tu día aún más especial, te regalamos un *10% DE DESCUENTO* en cualquier pedido que realices hoy. ¡Aprovecha esta oportunidad única! 💫
        ¡Que tengas un día increíble! 🥳✨"""

        clipboard = QApplication.clipboard()
        clipboard.setText(message)
        
        QMessageBox.information(
            self, 
            "Mensaje Copiado", 
            f"✅ Mensaje de felicitación para {client_name} copiado al portapapeles.\n\n"
        )
        
