import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QMessageBox, QGroupBox, QProgressBar, QFrame, QComboBox, QCheckBox
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QPixmap
from utils import http_get, http_put
from urls import API_URL_ORDERS
from datetime import datetime 
from openpyxl.utils import get_column_letter


class ExcelGenerationThread(QThread):
    progress = Signal(int)
    finished = Signal(str)
    error = Signal(str)

    def __init__(self, orders_data, semana, excel_mode, existing_file_path=None):
        super().__init__()
        self.orders_data = orders_data
        self.semana = semana
        self.excel_mode = excel_mode
        self.existing_file_path = existing_file_path

    def run(self):
        try:
            excel_data = []
            total_orders = len(self.orders_data)
            order_ids = []
            
            for i, order_data in enumerate(self.orders_data):
                if not order_data.get('added_to_excel', False):
                    order_ids.append(order_data['idOrder'])

                    progress = int((i + 1) / total_orders * 100)
                    self.progress.emit(progress)
                    
                    for book_info in order_data['books']:
                        book = book_info['book']
                        additives = book_info['additives']
                        
                        tipo = "R"
                        for additive in additives:
                            if additive["name"].lower().startswith("servicio"):
                                tipo = additive["name"][9].upper()
                                break
                        
                        portada = "normal"
                        tipo_caratula = "normal" 
                        
                        for additive in additives:
                            additive_name = additive["name"].lower()
                            if additive_name.startswith("carátula"):
                                portada = additive["name"][9:]
                                if "dura" in additive_name:
                                    if "premium" in additive_name:
                                        tipo_caratula = "dura_premium"
                                    else:
                                        tipo_caratula = "dura"
                                elif "solapa" in additive_name:
                                    tipo_caratula = "solapa"
                                else:
                                    tipo_caratula = "normal"
                                break
                        
                        def get_lomo(paginas, caratula_type):
                            if caratula_type in ["dura", "dura_premium"]:
                                return round(paginas / 170 + 0.5, 2)
                            else:
                                return round(paginas / 170 + 0.1, 4)
                        
                        numero_paginas = book.get('number_pages', 0)
                        lomo_calculado = get_lomo(numero_paginas, tipo_caratula)
                        
                        row_data = {
                            "Semana": self.semana,
                            "Tipo": tipo,
                            "Orden": order_data['idOrder'],
                            "Libro": book.get('title', 'Desconocido'),
                            "Paginas": numero_paginas,
                            "Cant": book_info['quantity'],
                            "Portada": portada,
                            "Venta": order_data['total_price'],
                            "Impreso": None, 
                            "Caratula": None, 
                            "Pegado": None,   
                            "Listo": None,    
                            "Entregado": None, 
                            "Lomo": lomo_calculado,
                        }
                        excel_data.append(row_data)
            
            if not excel_data:
                self.error.emit("No hay órdenes nuevas para añadir al Excel.")
                return

            df = pd.DataFrame(excel_data)
            print("🔀 Ordenando datos por Semana y Número de Orden...")
            df = df.sort_values(by=['Semana', 'Orden'], ascending=[True, True])
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            
            if self.excel_mode == 'new':
                filename = f"LOTE {datetime.now().month}.xlsx"
                file_path = os.path.join(desktop_path, filename)               
                self._save_excel_with_styles(df, file_path)
            else:
                file_path = self.existing_file_path
                self._append_to_existing_excel(df, file_path)

            self.finished.emit(file_path)
        except Exception as e:
            self.error.emit(str(e))

    def _save_excel_with_styles(self, df, file_path):
        wb = Workbook()



        ws = wb.active
        ws.title = str(datetime.now().month)
        headers = list(df.columns)
        ws.append(headers)
        for _, row in df.iterrows():
            ws.append(row.tolist())   

        self._add_footer_info(ws, df) 
        self._apply_excel_styles(ws, df)

        ws_gastos = wb.create_sheet("gastos")
        # self._setup_gastos_sheet(ws_gastos)

        ws_resumen = wb.create_sheet("resumen")
        # self._setup_resumen_sheet(ws_resumen)
    
        wb.save(file_path)
        
        print(f"✅ Excel NUEVO guardado en: {file_path}")
        print(f"📊 Filas totales: {ws.max_row}")

    def _append_to_existing_excel(self, df, file_path):
        print(f"📁 Intentando añadir a archivo existente: {file_path}")
        
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            print(f"📂 Archivo existente cargado. Filas actuales: {ws.max_row}")

            existing_data = []
            headers = [cell.value for cell in ws[1]]

            data_end_row = 1
            for row in range(2, ws.max_row + 1):
                cell_value = ws[f'A{row}'].value
                if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
                    break
                if cell_value is not None:
                    data_end_row = row
                    existing_row = []
                    for col in range(1, len(headers) + 1):
                        existing_row.append(ws.cell(row=row, column=col).value)
                    existing_data.append(existing_row)

            print(f"📊 Datos existentes encontrados: {len(existing_data)} filas (hasta fila {data_end_row})")

            ws.delete_rows(1, ws.max_row)
            print("🗑️  Todo el contenido del Excel eliminado")

            if existing_data:
                existing_df = pd.DataFrame(existing_data, columns=headers)
                combined_df = pd.concat([existing_df, df], ignore_index=True)
            else:
                combined_df = df

            print("🔀 Reordenando todos los datos...")
            combined_df = combined_df.sort_values(by=['Semana', 'Orden'], ascending=[True, True])
            print(f"✅ Datos combinados ordenados: {len(combined_df)} filas totales")

            ws.append(headers)
            for _, row in combined_df.iterrows():
                ws.append(row.tolist())
            
            print(f"📝 Todos los datos escritos: {len(combined_df)} filas")

            self._add_footer_info(ws, combined_df)
            
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Órdenes"
            print("🆕 Nuevo archivo creado")
            headers = list(df.columns)
            ws.append(headers)
            for _, row in df.iterrows():
                ws.append(row.tolist())
            print(f"✅ Filas añadidas al nuevo archivo: {len(df)}")
            self._add_footer_info(ws, df)
        
        self._apply_excel_styles(ws, df if 'combined_df' not in locals() else combined_df)

        wb.save(file_path)
        
        print(f"💾 Archivo guardado exitosamente con todos los datos ordenados")

    def _add_footer_info(self, ws, df):
        print("📝 Añadiendo fórmulas de Excel al final...")
        for _ in range(5):
            ws.append([None] * len(df.columns))
        
        last_data_row = len(df) + 1
        
        footer_row = [
            f"=COUNTA(A2:A{last_data_row})",
            "",
            f"=IF(COUNT(C2:C{last_data_row})>=2,MAX(C2:C{last_data_row})-MIN(C2:C{last_data_row}),0)",  # C: max-min con validación
            f"=COUNTA(A2:A{last_data_row})",
            f"=SUM(E2:E{last_data_row})",
            "",
            f"=COUNTA(A2:A{last_data_row})",
            f"=SUM(H2:H{last_data_row})",
            f'=COUNTIF(I2:I{last_data_row},"<>")',
            f'=COUNTIF(J2:J{last_data_row},"<>")',
            f'=COUNTIF(K2:K{last_data_row},"<>")',
            f'=COUNTIF(L2:L{last_data_row},"<>")',
            f'=COUNTIF(M2:M{last_data_row},"<>")',
            ""
        ]
        
        ws.append(footer_row)

        footer_formula_row = ws.max_row
        porcentaje_row = [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            f'=IF($A${footer_formula_row}>0,I{footer_formula_row}/$A${footer_formula_row},0)',
            f'=IF($A${footer_formula_row}>0,J{footer_formula_row}/$A${footer_formula_row},0)',
            f'=IF($A${footer_formula_row}>0,K{footer_formula_row}/$A${footer_formula_row},0)',
            f'=IF($A${footer_formula_row}>0,L{footer_formula_row}/$A${footer_formula_row},0)',
            f'=IF($A${footer_formula_row}>0,M{footer_formula_row}/$A${footer_formula_row},0)',
            ""
        ]
        
        ws.append(porcentaje_row)
        print(f"✅ Fórmulas de Excel añadidas al final")
        print(f"   📊 Referencia de datos: filas 2 a {last_data_row}")
        print(f"   📏 Footer en filas: {footer_formula_row} (fórmulas) y {footer_formula_row + 1} (porcentajes)")

    def _apply_excel_styles(self, ws, df):
        if ws.max_row == 0:
            print("⚠️  No hay datos para aplicar estilos")
            return
            
        print(f"🎨 Aplicando estilos a {ws.max_row} filas...")

        header_font = Font(bold=True, color="000000", size=12)
        header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        light_gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        footer_font = Font(bold=True, size=11)
        footer_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        porcentaje_font = Font(italic=True, size=10)
        porcentaje_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        headers = list(df.columns)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        last_data_row = len(df) + 1  
        footer_start_row = last_data_row + 6 
        footer_formula_row = footer_start_row
        footer_porcentaje_row = footer_start_row + 1

        print(f"🎯 Estructura del archivo:")
        print(f"   - Encabezado: fila 1")
        print(f"   - Datos: filas 2 a {last_data_row} ({len(df)} órdenes)")
        print(f"   - Líneas vacías: filas {last_data_row + 1} a {footer_start_row - 1}")
        print(f"   - Footer fórmulas: fila {footer_formula_row}")
        print(f"   - Footer porcentajes: fila {footer_porcentaje_row}")

        for row_idx in range(2, ws.max_row + 1):
            if row_idx == footer_porcentaje_row:
                row_fill = porcentaje_fill
                current_font = porcentaje_font
                row_type = "PORCENTAJE"
            elif row_idx == footer_formula_row:
                row_fill = footer_fill
                current_font = footer_font
                row_type = "FÓRMULA"
            elif last_data_row < row_idx < footer_start_row:
                row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                current_font = None
                row_type = "VACÍA"
            else:
                row_fill = white_fill if row_idx % 2 == 0 else light_gray_fill
                current_font = None
                row_type = "DATOS"
            
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.fill = row_fill
                cell.border = thin_border
                
                if current_font:
                    cell.font = current_font

                if row_type == "PORCENTAJE" and col_num in [9, 10, 11, 12, 13]:
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif row_type == "FÓRMULA":
                    if col_num == 8:
                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_num in [1, 3, 4, 5, 7, 9, 10, 11, 12, 13]:  # Números
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in [1, 3, 5, 6, 14]: 
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in [9, 10, 11, 12, 13]:  
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num == 8:  
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '"$"#,##0.00'
                elif col_num == 14:  
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = "0.00"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        print("📏 Ajustando anchos de columna...")

        fixed_widths = {
            'A': 8,
            'B': 8,
            'C': 12,
            'E': 10,
            'F': 8,
            'G': 15,
            'H': 12,
            'I': 10,
            'J': 10,
            'K': 10,
            'L': 8,
            'M': 12,
            'N': 8
        }

        for col, width in fixed_widths.items():
            ws.column_dimensions[col].width = width
            print(f"   📐 Columna {col}: {width} (ancho fijo)")

            print("   📚 Ajustando automáticamente columna D (Libro)...")
            column_d = ws['D']
            max_length = 0

            for cell in column_d:
                try:
                    if cell.value:
                        cell_value = str(cell.value)
                        if '\n' in cell_value:
                            lines = cell_value.split('\n')
                            max_line_length = max(len(line) for line in lines)
                            cell_length = max_line_length
                        else:
                            cell_length = len(cell_value)
                        
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 80) 
            adjusted_width = max(adjusted_width, 20)

            ws.column_dimensions['D'].width = adjusted_width
            print(f"   📐 Columna D: {adjusted_width} (longitud máxima: {max_length})")
                
        if last_data_row > 1:
            try:
                ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{last_data_row}"
                print("🔍 Filtros aplicados solo a datos")
            except Exception as e:
                print(f"⚠️  Error aplicando auto_filter: {e}")

        if ws.max_row >= 2:
            ws.freeze_panes = "A2"
            print("❄️  Paneles congelados")
        
        print("✅ Estilos aplicados correctamente")

    def _apply_excel_styles_to_new_rows(self, ws, df, start_row):
        if start_row > ws.max_row:
            print("⚠️  No hay nuevas filas para aplicar estilos")
            return
            
        print(f"🎨 Aplicando estilos a nuevas filas desde {start_row} hasta {ws.max_row}")
        
        light_gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        headers = list(df.columns)
        
        for row_idx in range(start_row, ws.max_row + 1):
            row_fill = white_fill if row_idx % 2 == 0 else light_gray_fill
            
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.fill = row_fill
                cell.border = thin_border
                
                if col_num in [1, 3, 5, 6, 14]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in [9, 10, 11, 12, 13]: 
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num == 8:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '"$"#,##0.00'
                elif col_num == 14: 
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = "0.00"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        print("✅ Estilos aplicados a nuevas filas")

    def _mark_orders_as_added_to_excel(self, order_ids):
        try:
            print(f"🏷️  Marcando {len(order_ids)} órdenes como añadidas al Excel...")
            for order_id in order_ids:
                update_data = {"added_to_excel": True}
                response = http_put(f"{API_URL_ORDERS}{order_id}/update_order_data/", update_data)
                if response and response.status_code == 200:
                    print(f"✅ Orden {order_id} marcada como añadida al Excel")
                else:
                    error_msg = response.text if response else "Sin respuesta"
                    print(f"❌ Error marcando orden {order_id}: {error_msg}")
        except Exception as e:
            print(f"❌ Error al marcar órdenes como añadidas al Excel: {e}")


class ExcelTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_excel_tab()

    def _setup_excel_tab(self):
        main_layout = QVBoxLayout(self)
        main_layout.setAlignment(Qt.AlignTop)

        #? ---------------- ENCABEZADO PRINCIPAL ----------------
        header_layout = QHBoxLayout()
        header_icon = QLabel()
        header_icon.setPixmap(
            QPixmap("icons/table.png").scaled(50, 50, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        )
        header_title = QLabel("Generar Reporte Excel")
        header_title.setStyleSheet("""
            font-size: 28px;
            font-weight: bold;
            color: #333;
        """)
        header_layout.addWidget(header_icon)
        header_layout.addWidget(header_title)
        header_layout.addStretch()

        line = QLabel()
        line.setFixedHeight(2)
        line.setStyleSheet("background-color: #d0d0d0; margin: 8px 0;")

        main_layout.addLayout(header_layout)
        main_layout.addWidget(line)

        #? ---------------- INSTRUCCIONES ----------------
        instructions_group = QGroupBox("Instrucciones")
        instructions_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #2E86C1;
                border-radius: 8px;
                padding: 10px;
                background-color: #f0f8ff;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #2E86C1;
            }
        """)
        
        instructions_layout = QVBoxLayout(instructions_group)
        instructions_label = QLabel(
            "Esta herramienta generará un archivo Excel con las órdenes del sistema.\n\n"
            "• <b>Continuar Excel existente:</b> Añade órdenes nuevas al archivo actual\n"
            "• <b>Nuevo Excel mensual:</b> Crea un nuevo archivo LOTE_MM_mes.xlsx\n\n"
            "Solo se incluirán órdenes que no han sido añadidas anteriormente al Excel.\n\n"
            "📏 <b>Cálculo automático del Lomo:</b>\n"
            "• Carátula Normal o con Solapa: Páginas/170 + 0.1\n"
            "• Carátula Dura: Páginas/170 + 0.5\n\n"
            "✅ <b>Checkboxes interactivos:</b> Las columnas Impreso, Carátula, Pegado, Listo y Entregado\n"
            "tendrán checkboxes reales que pueden marcarse/hacerse clic en Excel."
        )
        instructions_label.setWordWrap(True)
        instructions_label.setStyleSheet("font-size: 12px; color: #555;")
        instructions_layout.addWidget(instructions_label)
        main_layout.addWidget(instructions_group)

        #? ---------------- CONTROLES ----------------
        controls_group = QGroupBox("Generar Reporte")
        controls_layout = QVBoxLayout(controls_group)

        semana_layout = QHBoxLayout()
        semana_label = QLabel("Seleccionar semana:")
        semana_label.setStyleSheet("font-weight: bold;")
        self.semana_combo = QComboBox()
        self.semana_combo.addItems(["1", "2", "3", "4"])
        self.semana_combo.setCurrentIndex(0)
        self.semana_combo.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 1.5px solid #ccc;
                border-radius: 6px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:focus {
                border: 1.5px solid #2E86C1;
            }
        """)
        semana_layout.addWidget(semana_label)
        semana_layout.addWidget(self.semana_combo)
        semana_layout.addStretch()
        controls_layout.addLayout(semana_layout)

        excel_mode_layout = QHBoxLayout()
        excel_mode_label = QLabel("Modo de Excel:")
        excel_mode_label.setStyleSheet("font-weight: bold;")
        
        self.excel_mode_combo = QComboBox()
        self.excel_mode_combo.addItems(["Continuar Excel existente", "Nuevo Excel mensual"])
        self.excel_mode_combo.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 1.5px solid #ccc;
                border-radius: 6px;
                font-size: 14px;
                background-color: white;
            }
        """)
        
        excel_mode_layout.addWidget(excel_mode_label)
        excel_mode_layout.addWidget(self.excel_mode_combo)
        excel_mode_layout.addStretch()
        controls_layout.addLayout(excel_mode_layout)

        columns_info = QLabel(
            "Columnas del reporte:\n"
            "• Semana, Tipo, Orden, Libro, Páginas, Cantidad, Portada\n"
            "• Venta, Impreso, Carátula, Pegado, Listo, Entregado, Lomo\n\n"
            "📏 <b>Lomo calculado automáticamente:</b>\n"
            "• Normal/Solapa: Páginas/170 + 0.1\n"
            "• Dura: Páginas/170 + 0.5\n\n"
            "✅ Las columnas de estado tendrán checkboxes interactivos en Excel"
        )
        columns_info.setWordWrap(True)
        columns_info.setStyleSheet("font-size: 11px; color: #666; padding: 8px; background-color: #f9f9f9; border-radius: 5px;")
        controls_layout.addWidget(columns_info)

        self.generate_btn = QPushButton("📊 Generar Reporte Excel")
        self.generate_btn.setObjectName("primaryBtn")
        self.generate_btn.clicked.connect(self._generate_excel)
        controls_layout.addWidget(self.generate_btn, alignment=Qt.AlignCenter)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #ccc;
                border-radius: 4px;
                text-align: center;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #2E86C1;
                border-radius: 3px;
            }
        """)
        controls_layout.addWidget(self.progress_bar)

        main_layout.addWidget(controls_group)
        main_layout.addStretch()

    def _generate_excel(self):
        print("🚀 Iniciando generación de Excel...")
        self.generate_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)

        semana = int(self.semana_combo.currentText())
        excel_mode = 'append' if self.excel_mode_combo.currentText() == "Continuar Excel existente" else 'new'

        print(f"📋 Parámetros - Semana: {semana}, Modo: {excel_mode}")

        existing_file_path = None
        if excel_mode == 'append':
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            existing_files = [f for f in os.listdir(desktop_path) if f.startswith(f"LOTE {datetime.now().month}") and f.endswith(".xlsx")]
            if existing_files:
                existing_files.sort(key=lambda x: os.path.getctime(os.path.join(desktop_path, x)), reverse=True)
                existing_file_path = os.path.join(desktop_path, existing_files[0])
                print(f"📂 Archivo existente encontrado: {existing_file_path}")
            else:
                QMessageBox.warning(self, "Advertencia", 
                                  "No se encontró un archivo Excel existente. Se creará uno nuevo.")
                excel_mode = 'new'
                print("⚠️  No se encontró archivo existente, creando nuevo")

        try:
            print("🌐 Obteniendo órdenes del servidor...")
            response = http_get(API_URL_ORDERS)
            
            if not response or response.status_code != 200:
                QMessageBox.warning(self, "Error", "No se pudieron obtener las órdenes del sistema.")
                self._reset_ui()
                return
            orders_list = response.json()
            print(f"📦 Total de órdenes obtenidas: {len(orders_list)}")
            
            if not orders_list:
                QMessageBox.information(self, "Información", "No hay órdenes en el sistema.")
                self._reset_ui()
                return

            orders_data = []
            total_orders = len(orders_list)
            
            progress_label = QLabel("Obteniendo datos de órdenes...")
            self.layout().addWidget(progress_label)
            
            for i, order in enumerate(orders_list):
                order_id = order.get('idOrder')
                if order_id:
                    order_data = self._get_order_data(order_id)
                    if order_data:
                        orders_data.append(order_data)
                
                progress = int((i + 1) / total_orders * 50)
                self.progress_bar.setValue(progress)
            
            progress_label.deleteLater()
            
            if not orders_data:
                QMessageBox.warning(self, "Error", "No se pudieron obtener datos de ninguna orden.")
                self._reset_ui()
                return

            new_orders = [order for order in orders_data if not order.get('added_to_excel', False)]
            
            print(f"🆕 Órdenes nuevas (no añadidas al Excel): {len(new_orders)}")
            
            if not new_orders:
                QMessageBox.information(self, "Información", 
                                      "No hay órdenes nuevas para añadir al Excel.\n"
                                      "Todas las órdenes ya han sido procesadas anteriormente.")
                self._reset_ui()
                return

            print("🔄 Iniciando thread de generación de Excel...")
            self.thread = ExcelGenerationThread(new_orders, semana, excel_mode, existing_file_path)
            self.thread.progress.connect(self._update_progress)
            self.thread.finished.connect(self._on_excel_generated)
            self.thread.error.connect(self._on_excel_error)
            self.thread.start()

        except Exception as e:
            print(f"❌ Error crítico: {e}")
            QMessageBox.critical(self, "Error", f"Error al procesar órdenes: {str(e)}")
            self._reset_ui()

    def _get_order_data(self, order_id):
        url = f"{API_URL_ORDERS}{order_id}/full_details/"
        r = http_get(url)
        if r and r.status_code == 200:
            return r.json()
        return None

    def _update_progress(self, value):
        adjusted_value = 50 + int(value * 0.5)
        self.progress_bar.setValue(adjusted_value)

    def _on_excel_generated(self, file_path):
        print(f"✅ Proceso completado. Archivo: {file_path}")
        self._reset_ui()
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            order_count = ws.max_row - 1 
            print(f"📊 Archivo contiene {order_count} órdenes")
        except Exception as e:
            order_count = "desconocido"
            print(f"⚠️  Error contando órdenes: {e}")
        
        mode_text = "añadidas al Excel existente" if self.excel_mode_combo.currentText() == "Continuar Excel existente" else "en nuevo Excel mensual"
        
        QMessageBox.information(
            self, 
            "Éxito", 
            f"Reporte Excel generado exitosamente:\n{file_path}\n\n"
            f"Semana: {self.semana_combo.currentText()}\n"
            f"Órdenes procesadas: {order_count}\n"
            f"Modo: {mode_text}\n\n"
            f"📏 <b>Lomo calculado automáticamente:</b>\n"
            f"• Carátula Normal/Solapa: Páginas/170 + 0.1\n"
            f"• Carátula Dura: Páginas/170 + 0.5\n\n"
            f"✅ <b>Checkboxes interactivos:</b> Las columnas Impreso, Carátula, Pegado, Listo y Entregado\n"
            f"tienen checkboxes reales que pueden marcarse/hacerse clic en Excel."
        )

    def _on_excel_error(self, error_message):
        print(f"❌ Error en generación de Excel: {error_message}")
        self._reset_ui()
        QMessageBox.critical(self, "Error", f"No se pudo generar el reporte Excel: {error_message}")

    def _reset_ui(self):
        self.generate_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.progress_bar.setValue(0)

    def _apply_styles(self):
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #b0b0b0;
                border-radius: 8px;
                padding: 12px;
                background-color: #fff;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #333;
            }
            QPushButton#primaryBtn {
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #333, stop:1 #000);
                color: white;
                font-weight: 600;
                padding: 10px 20px;
                border-radius: 8px;
                border: 1px solid #111;
            }
            QPushButton#primaryBtn:hover {
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #444, stop:1 #111);
                border: 1px solid #222;
            }
            QPushButton#primaryBtn:pressed {
                background-color: #000;
                border: 1px solid #000;
                padding-top: 11px;  
                padding-bottom: 9px;
            }
        """)